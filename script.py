import time
import pyautogui
import openpyxl
import threading
import keyboard
from datetime import datetime
from tkinter import *
from cryptography.fernet import Fernet
from selenium.webdriver.common.keys import Keys
from tkinter import filedialog



#automation
dateToGo = ''
##########################################################################################################
def sleeping(z):
    time.sleep(int(z.strip()))

def click(clicks,z):
    if(clicks == 1):
        pyautogui.click(button=z)
    else:
        pyautogui.doubleClick(button=z)

def sendKeys(z):
    pyautogui.typewrite(z)

def moveTo(x,y):
    pyautogui.moveTo(x=x,y=y)

def waitUntilFound(rgb,waitTime):
    waitTime = waitTime / 2 
    check = True
    counter = 0
    while(check):
        check = pyautogui.pixelMatchesColor(86, 123,rgb)
        time.sleep(2)
        if(counter > waitTime):
            print('waited for too long exception')
            return False
        counter += 1
    return True
    

def dragAndDrop(x1,y1,x2,y2):
    pyautogui.moveTo(x=int(x1.strip()),y=int(y1.strip()))
    pyautogui.dragTo(x=int(x2.strip()),y=int(y2.strip()),button='left')
    
def holdAndMove(x1, y1, x2, y2):
    pyautogui.mouseDown(button='left', x=x1, y=y1)
    pyautogui.moveTo(x2, y2)
    pyautogui.mouseUp(button='left')

def is_dialogue_box_present(dialogue_box_image):
    dialogue_box_position = pyautogui.locateOnScreen(dialogue_box_image)
    if dialogue_box_position is not None:
        return True
    else:
        return False
    


#excelWork
#########################################################################################################################
listOfQuery = []

def parserStr(stri):
    numberFromStr = ''
    for i in stri:
        if(i.isdigit()):
            numberFromStr += i
    return numberFromStr

def getDataFromExcel(filePath):
    wb = openpyxl.load_workbook(filePath)
    sheet = wb.active
    j = int(tagNumber.get())
    for i in range(2,sheet.max_row+1):
        latestTime.set(f"done {i} out of {sheet.max_row}")
        temp = []
        temp.append(parserStr(str(sheet.cell(row=i,column=j).value)))
        listOfQuery.append(temp)



def loadExcelFile():
     global filePath
     latestTime.set("Wait Loading....")
     load_button.config(state=DISABLED,bg='red')
     filePath = filedialog.askopenfilename()
     global listOfQuery
     getDataFromExcel(filePath)
     var = "Last Updated: " + str(time.strftime("%H:%M:%S", time.localtime()))
     latestTime.set(var)
     load_button.config(state=ACTIVE,bg=orig_color)

def loadTheFile():
    t = threading.Thread(target=loadExcelFile)
    t.start()
    return

#uiwork     
#########################################################################################################################

root = Tk()
root.title("Auto_Survey")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.minsize(650,600)
root.geometry("650x600")
root.maxsize(650,600)
root.resizable(width=False, height=False)  

#load button
load_button = Button(root, text="LOAD EXCEL FILE", command=loadTheFile)
load_button.grid(row=1,column=1,padx=0,pady=0)
orig_color = load_button.cget("background")

latestTime = StringVar()
latestTime.set("Please Load The File")
modifiedLabel = Label(root,textvariable=latestTime).grid(row=1,column=0,padx=0)


userNameLabel = Label(root, text="enter the auth name:")
userNameLabel.grid(row=0, column=0)

userName = Entry(root)
userName.grid(row=0, column=1)


tagNumberLabel = Label(root, text="tag number column:")
tagNumberLabel.grid(row=0, column=2, padx=5, pady=5, sticky=W)

tagNumber = Entry(root)
tagNumber.grid(row=0, column=3, padx=5, pady=5)

dateRangeLabel = Label(root, text="enter ranged base date:")
dateRangeLabel.grid(row=2, column=0, padx=5, pady=5, sticky=W)

dateRange = Text(root, width=20, height=10)
dateRange.grid(row=3, column=0, columnspan=4, padx=50, pady=0)
##############################################################

failureLabel = Label(root, text="failure:")
failureLabel.grid(row=6, column=1, padx=5, pady=5, sticky=W)

failureList = Text(root, width=20, height=15)
failureList.grid(row=7, column=0, columnspan=4, padx=5, pady=0)

successLabel = Label(root, text="success:")
successLabel.grid(row=6, column=3, padx=5, pady=5, sticky=W)

successList = Text(root, width=20, height=15)
successList.grid(row=7, column=3, columnspan=4, padx=5, pady=0)


def on_change1(event):
    textOfOutputArea = tagNumber.get()
    with open ("tagNumber.txt","w") as f:
        f.write(textOfOutputArea)
    
def on_change2(event):
    textOfOutputArea = userName.get()
    with open ("userName.txt","w") as f:
        f.write(textOfOutputArea)


def on_modified1(event):
    textOfOutputArea = dateRange.get('1.0', 'end')
    with open ("date.txt","w") as f:
        f.write(textOfOutputArea)

def on_modified2(event):
    textOfOutputArea = successList.get('1.0', 'end')
    with open ("success.txt","w") as f:
        f.write(textOfOutputArea)

def on_modified3(event):
    textOfOutputArea = failureList.get('1.0', 'end')
    with open ("failure.txt","w") as f:
        f.write(textOfOutputArea)

tagNumber.bind('<KeyRelease>', on_change1)
userName.bind('<KeyRelease>', on_change2)
dateRange.bind('<KeyRelease>', on_modified1)
successList.bind('<KeyRelease>', on_modified2)
failureList.bind('<KeyRelease>', on_modified3)

def fillOutputArea(outputBox,query,mode):
    if(mode == 0):
        outputBox.insert(END,query+'\n')
    else :
        outputBox.insert(END,query)


with open ("date.txt","r") as f:
    for line in f:
        fillOutputArea(dateRange,line,1)

with open ("success.txt","r") as f:
    for line in f:
        fillOutputArea(successList,line,1)

with open ("failure.txt","r") as f:
    for line in f:
        fillOutputArea(failureList,line,1)

with open ("tagNumber.txt","r") as f:
    for line in f:
        tagNumber.insert(0,line)

with open ("userName.txt","r") as f:
    for line in f:
        userName.insert(0,line)

    



#programLogic
########################################################################################################################
itr = 0
def setDate(z):
    today = str(datetime.now())
    today = today.split(' ')
    today = today[0].split('-')
    userDate = int(z.split('-')[0])
    userMonth = int(z.split('-')[1])
    todayDate = int(today[2])
    todayMonth = int(today[1])

    pyautogui.moveTo(592,310)
    pyautogui.click()

    for i in range(0,todayMonth-userMonth):
        pyautogui.moveTo(404,332)
        pyautogui.click()

    if(todayMonth == userMonth):
        for i in range(0,todayDate-userDate):
            pyautogui.press('left')
    else:
        for i in range(0,userDate-1):
            pyautogui.press('right')
    pyautogui.press('enter')


dateList = []
def fillDateList():
     listOfString = dateRange.get("1.0",END).split('\n')
     for i in range(0,len(listOfString)):
         if(len(listOfString[i]) < 10): continue
         temp = listOfString[i].strip()
         a = temp.split(':')[0]
         b = temp.split(':')[1]
         c = temp.split(':')[2]
         for i in range(int(a)-1,int(b)):
            dateList.append([int(a),int(b),c])
######################################################################
def decrypt_string(encrypted_string):
    key = 'ECPHuqGMo6QE2tcLElUX2GBmvOngpzFTbPAO09KMqdo='
    f = Fernet(key)
    decrypted = f.decrypt(encrypted_string)
    return decrypted.decode()

authenticator = False
def checkTheauth():
    global authenticator
    with open("allowed_users.txt","r") as f:
        for line in f:
            auth = line
            auth = decrypt_string(str(auth).strip())
            temp = str(userName.get()).split()
            temp = temp[0]
            if(temp in auth):
                authenticator = True
                return
######################################################################
def logic(dictCheck):
    global dateToGo
    global itr
    j = int(tagNumber.get().strip())
    if(dictCheck.get(listOfQuery[itr][0]) == 1):
        itr+=1
        return
    state.set('RUNNING')
    z = 7
    for i in range(0,z):
        time.sleep(1)
        warnLogger.set("minimize the program in " + str(z-i) + " seconds")
    warnLogger.set("program is running")

    pyautogui.press('tab')
    
    moveTo(699,117)
    click(2,'left')
    pyautogui.typewrite('\b')
    sendKeys(listOfQuery[itr][0])

    moveTo(831,114)
    click(1,'left')
    time.sleep(3)

    moveTo(224,224)
    click(1,'left')
    time.sleep(1)

    moveTo(489,332)
    click(1,'left')

    moveTo(489,350)
    click(1,'left')

    setDate(dateToGo)

    moveTo(875,462)
    click(1,'left')
    time.sleep(1)

    moveTo(531,376)
    rgb = (8,131,216)
    check = waitUntilFound(rgb,5)

    with open ("success.txt","a") as f:
        f.write(listOfQuery[itr][0]+'\n')

    if(check == True):
        moveTo(748,435)
        click(1,'left')
            
        time.sleep(3)

        moveTo(748,435)
        click(1,'left')

    moveTo(830,435)
    click(1,'left')
    itr+=1


    



    




    
#start button
########################################################################################################################
def preLogic():
    fillDateList()
    global itr
    global dateToGo
    alreadyDone = []
    checkTheauth()
    if(authenticator == False):
        warnLogger.set("not allowed")
        return
    warnLogger.set("allowed user")
    alreadyDone = successList.get("1.0",END).split('\n')
    dictCheck = {}
    for i in range(0,len(alreadyDone)):
        dictCheck[alreadyDone[i].strip()] = 1
    i = 0
    while(i<len(dateList)):
        itr = dateList[i][0] - 1
        dateToGo = dateList[i][2]
        while(itr < dateList[i][1]):
            logic(dictCheck)
        i+=1
    warnLogger.set("completed")

def startProgram():
    startButton.config(state=DISABLED,bg='LIGHT GREEN')
    for i in range(0,1):    
        threading.Thread(target=preLogic).start()
    startButton.config(state=NORMAL,bg=orig_color)

state = StringVar()
state.set('START')
startButton = Button(root, textvariable=state, command=startProgram)
startButton.grid(row=4,column=0,padx=0,pady=0)

warnLogger = StringVar()
warnLogger.set("click on start to run the program :")
warnLabel = Label(root,textvariable=warnLogger)
warnLabel.grid(row=5,column=0,padx=0)

#mainFunction
#######################################################################################################################
def stop_program():
    root.quit()

def check_hotkey():
    if keyboard.is_pressed('ctrl') and keyboard.is_pressed('q'):
        stop_program()
    else:
        root.after(100, check_hotkey)


root.after(100, check_hotkey)
#######################################################################################################################
root.mainloop()


 